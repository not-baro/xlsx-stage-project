from openpyxl import load_workbook
from flask import Flask, request, render_template, redirect, url_for, send_from_directory, send_file
import time
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import xlsxwriter
import re
import pandas
import zipfile
import io

app = Flask(__name__)

@app.route('/')
def index():
        return render_template('index.html')

@app.route('/modify')
def modify():
    return render_template('uploader.html')

@app.route('/output', methods = ['GET', 'POST'])
def upload():
    if request.method == 'POST':
        f = request.files['file']
        f.save(f.filename)
        return process(f.filename)

def newSubstring(listacit,i):
    newSub = (listacit[i][0:30])
    newSub = newSub.strip()
    newSub = newSub.replace(',','').replace('é','è').replace('ù','u').replace('"','').replace("'",'').replace('‘','').replace('’','').replace('-','').replace('—',' ').replace('[...]','').replace('[....]','').replace('(','').replace(')','').replace('e.g.','').replace('[','').replace(']','').replace('  ',' ').replace('”','').replace('  ','').replace('   ','').replace('[1]','').replace('[2]','').replace('[3]','').replace('[4]','').replace('[5]','').replace('[6]','').replace('[7]','').replace('[8]','').replace('[9]','').replace('[10]','').replace('6.','').replace('1.','').replace('2.','').replace('3.','').replace('4.','').replace('5.','').replace('7.','').replace('8.','').replace('9.','').replace('10.','').replace('.experience','').replace('.first','').replace('.tori','').replace('.6','').replace('i.e.','').replace('I.e.','').replace('I.E.','').replace('.first.','first.').replace(' s.','').replace('“','').replace('work.oriented','work oriented')


    newSubSplit = newSub.split()

    if(len(newSubSplit)>5):
        newSubSplit.pop()
    return newSubSplit



def replaceString(listacit,i):
    newSub = (listacit[i])
    newSub = newSub.strip()
    newSub = newSub.replace(',','').replace('é','è').replace('ù','u').replace('"','').replace("'",'').replace('‘','').replace('’','').replace('-','').replace('—',' ').replace('[...]','').replace('[....]','').replace('(','').replace(')','').replace('e.g.','').replace('[','').replace(']','').replace('  ',' ').replace('”','').replace('  ','').replace('   ','').replace('[1]','').replace('[2]','').replace('[3]','').replace('[4]','').replace('[5]','').replace('[6]','').replace('[7]','').replace('[8]','').replace('[9]','').replace('[10]','').replace('6.','').replace('1.','').replace('2.','').replace('3.','').replace('4.','').replace('5.','').replace('7.','').replace('8.','').replace('9.','').replace('10.','').replace('.experience','').replace('.first','').replace('.tori','').replace('.6','').replace('i.e.','').replace('I.e.','').replace('I.E.','').replace('.first.','first.').replace(' s.','').replace('“','').replace('work.oriented','work oriented')

    return newSub

def newSubstring2(frase):
    newSub = (frase)
    newSub = newSub.strip()
    newSub = newSub.replace(',','').replace('é','è').replace('ù','u').replace('"','').replace("'",'').replace('‘','').replace('’','').replace('-','').replace('—','').replace('[...]','').replace('[....]','').replace('(','').replace(')','').replace('e.g.','').replace('[','').replace(']','').replace('  ',' ').replace('”','').replace('  ','').replace('   ','').replace('[1]','').replace('[2]','').replace('[3]','').replace('[4]','').replace('[5]','').replace('[6]','').replace('[7]','').replace('[8]','').replace('[9]','').replace('[10]','').replace('6.','').replace('1.','').replace('2.','').replace('3.','').replace('4.','').replace('5.','').replace('7.','').replace('8.','').replace('9.','').replace('10.','').replace('.experience','').replace('.first','').replace('.tori','').replace('.6','').replace('i.e.','').replace('I.e.','').replace('I.E.','').replace('.first.','first.').replace(' s.','')


    newSubSplit = newSub.split()

    if(len(newSubSplit)>5):
        newSubSplit.pop()
    return newSubSplit

#---fine funzioni utility

def process(uploaded):

    timestr = time.strftime("%d%m%Y")

    wb=load_workbook(uploaded)

    sheets = wb.sheetnames #trova i nomi degli sheet
    ws = wb.active
    writer = pandas.ExcelWriter(timestr + '.xlsx', engine='xlsxwriter')

    #extract colonne input
    tempo_A = ws['A']
    matricole_D = ws['D']
    doc_C = ws['C']
    citazioni_E = ws['E']
    mail_B = ws['B']
    #---------------------

    #set liste e dizionari
    lista_matr = []
    lista_cit = []
    lista_cit_2 = []
    lista_mail = []
    lista_tempo = []
    lista_doc = []
    lista_doc_and_prop = []
    sottofrasi=[]
    lista_k = []
    lista_z = []
    frasi_excel=[]

    doc_and_prop = {}
    prop_and_k = {}
    doc_and_cit = {}
    matrEmailDictionary = {}
    #-------------------

    #set variabili varie
    lunghezza = len(citazioni_E)
    p = lunghezza+4
    k=0
    z=0
    #-------------------

    #estrazione valori dalle colonne + le metto nelle liste corrisp.
    for i in range(1,lunghezza):
        lista_tempo.append(tempo_A[i].value)

    for i in range(1,lunghezza):
        if(matricole_D[i].value is None):
            matricole_D[i].value = 'non ins'
            citazioni_E[i].value = 'inserimento non valido'
        lista_matr.append(str(matricole_D[i].value).replace('.0',''))

    for i in range(1,lunghezza):
        if(doc_C[i].value is None):
            doc_C[i].value = 'doc. non ins'
        lista_doc.append(doc_C[i].value)

    for i in range(1,lunghezza):
        if(mail_B[i].value is None):
            mail_B[i].value = 'mail non valida/non ins.'
        lista_mail.append(mail_B[i].value)

    for i in range(1,lunghezza): #salta la prima casella del titolo se metto 1
        lista_cit.append(citazioni_E[i].value)

    #############
    #ricerca data ultimo inserimeno per persona
    #############
    v=2
    infocrono=''
    infocronolist=[]

    lista_mail_singole = []

    lista_mail_singole = list(dict.fromkeys(lista_mail)) #rimuove duplicati in lista mail

    #print(lista_mail_singole)

    for i in range(len(lista_mail_singole)):
        for j in range(1,lunghezza):
            if(ws.cell(row=v, column = 2).value == lista_mail_singole[i]):
                infocrono = ws.cell(row=v, column = 1).value
                #print(infocrono)
                #print('i = ' + str(i))
                #print('j = ' + str(j))
            v=v+1
        infocronolist.append(str(infocrono))
        v=2
    #print(infocrono)

#okkk!!
###############
###############

    ###############
    ###############
    #-----------------------------------------------------

    #selezione wb
    wbOutput = Workbook()
    ws = wbOutput.active
    #---------------------

    #divido in sottostringhe delimitate da un punto
    for i in range (len(lista_cit)):
        replaced = replaceString(lista_cit,i)
        splittato = (re.split('\.', replaced))

        #print(splittato)
        for h in range(len(splittato)):
            if(splittato[h].strip() != '' and splittato[h].strip() != ' ' and splittato[h].strip() != '  'and splittato[h].strip() != '   '):
                sottofrasi.append(splittato[h].strip())
                lista_doc_and_prop.append(lista_doc[i])
                lista_cit_2.append(lista_cit[i])
                #print(splittato[h].strip())

    #------------------------------------
    #rilevo ripetizioni

    for i in range(len(sottofrasi)):
        temp1 = newSubstring(sottofrasi,i) #estraggo prime parole della sottostringa e confronto con le altre
        s=' '

        for j in range (lunghezza):
            if(s.join(temp1) != '' and s.join(temp1) != ' ' and s.join(temp1) != '  '):
                if((s.join(temp1)) in sottofrasi[j]):

                    k=k+1
        #print('la stringa in posizione: ' + str(i) + ' appare ' + str(k) + ' volte')

        if(k==0):
            lista_k.append(k+1)
        else:
            lista_k.append(k)

        #print('resetto k')
        #print('-----------------------------')
        #print('')
        k=0


    #associazioni varie per non risultare sfalsate nella stampa finale
    #------------------------
    keys=sottofrasi
    values=lista_doc_and_prop
    doc_and_prop = list(zip(keys,values))
    #---------------------------

    keys2=sottofrasi
    values2=lista_k
    prop_and_k = list(zip(keys2,values2))
    #---------------------------

    #print('')
    #print('')

    #-----------------------------

    ws['B1'] = 'PROPOSIZIONE'
    ws['C1'] = 'FONTE'
    ws['A1'] = 'RIPETIZIONI'

    p=2
    f=0
    for i in range (len(prop_and_k)):
        #print(dictionary[i][0])
        #print(dictionary[i][1])
        ws.cell(row=p, column = 2).value = prop_and_k[i][0]
        ws.cell(row=p, column = 1).value = prop_and_k[i][1]
        ws.cell(row=p, column = 3).value = doc_and_prop[i][1]
        #ws.cell(row=p, column = 6).value = doc_and_cit[i][1] opzionale


        f=f+1

        p=p+1
    p=2

    lunghezzaOutput = len(ws['C'])

    wbOutput.save(filename="OutputDef4.xlsx")
    wb.close()

    #----------------------------------------
    #EDIT workaround - cancella righe con le proposizioni duplicate e le riordina

    print('entro nel new wb')
    df = pandas.read_excel('OutputDef4.xlsx')


    s=' '

    for x in range (len(lista_k)):
        frasi_excel.append(df.at[x, 'PROPOSIZIONE'])

    for i in range (lunghezzaOutput-1):
        if (df.at[i, 'RIPETIZIONI'] > 1):
            str1 = df.at[i, 'PROPOSIZIONE']
            for j in range (lunghezzaOutput - 1):
                str2=newSubstring(frasi_excel,j)
                if(s.join(str2) != '' and s.join(str2) != ' ' and s.join(str2) != '  '):
                    if (i != j and s.join(str2) in str1):
                        df.at[j, 'PROPOSIZIONE']=''
                        df.at[j, 'RIPETIZIONI']=None
                        df.at[j, 'FONTE']=''


    df=df.sort_values('RIPETIZIONI', ascending=False)
    #df.to_excel('OutputDef4.xlsx')

    #endEDIT
    #-----------------------------------------
    #stampo file matricole
    #conto matricole
    for x in range(lunghezza-1):
        tempMatr = (lista_matr[x])

        for i in range(lunghezza-1):

            if(tempMatr in lista_matr[i]):
                z=z+1


        #print('lo studente con matricola ' + lista_matr[x] + ' ha inserito ' + str(z) + ' citazioni')

        lista_z.append(z)
        z=0 #reset z

    keys2 = lista_matr
    values2 = lista_z
    matrDictionary = dict(zip(keys2, values2))

    #workaround

    wb1=Workbook()
    ws1=wb.active
    ws1.title='Matricole Sheet'
    wb1.save(filename = 'MatricoleDef4.xlsx')

    df2 = pandas.read_excel('MatricoleDef4.xlsx')

    df2.at[1, 1]='MATRICOLA'
    df2.at[1, 2]='EMAIL'
    df2.at[1, 3]='N.INSERIMENTI'
    df2.at[1, 4]='DATA ULTIMO INSERIMENTO'

    b=2

    #print(matrDictionary)
    print('')
    #print(matrDictionary.keys())


    for key in matrDictionary.keys():
        df2.at[b, 1] = key
        b=b+1

    b=2

    for values in matrDictionary.values():
        df2.at[b, 3] = values
        b=b+1

    b=2

    keys3 = lista_matr
    values3 = lista_mail
    matrEmailDictionary = dict(zip(keys3, values3))

    #print(matrEmailDictionary)
    for values in matrEmailDictionary.values():
        df2.at[b, 2] = values
        b=b+1
    b=2

    ##### scrittura dati crono e rimozione righe problematiche
    y=0
    for i in range (len(matrEmailDictionary)):

        if (df2.at[b,3] == 1):
            df2.at[b,1]=None
            df2.at[b,2]=None
            df2.at[b,3]=None
            df2.at[b,4]=None
            y=y-1
        else:
            df2.at[b, 4] = infocronolist[y]
        y=y+1
        print('y: ' + str(y))
        b=b+1

    b=2


    listaDuplicati = []
    for i in range (len(matrEmailDictionary)):
        listaDuplicati.append(df2.at[b,2])
        b=b+1
    b=2

    ################

    #print('len sottofrasi: ' + str(len(sottofrasi)))
    #print('len lista doc and prop: ' + str(len(lista_doc_and_prop)))
    #print('len doc and prop: ' + str(len(doc_and_prop)))
    #print('len lista k: ' + str(len(lista_k)))
    #print('len prop and k: ' + str(len(prop_and_k)))

    wb.close()
    #df2.to_excel('MatricoleDef4.xlsx')
    df.to_excel(writer, sheet_name='Citazioni')
    df2.to_excel(writer, sheet_name='Studenti')
    writer.save()

    directory = '/home/pytobaro/' + timestr + '.xlsx'
    return send_file(directory,'output.xlsx', as_attachment = True)



if __name__ == "__main__":
    app.run(debug=True)
